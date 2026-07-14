"""Parser for the Artech House transmittal workbook.

Sheet 1 holds labeled fields (Title:, Author(s):, ISBN:, Trim size:, Color:, ...) plus a
per-chapter word-count table; Sheet 2 holds the actual front-matter/copyright-page boilerplate
text. In practice PMs often leave Sheet 1's labeled value cells blank and only fill in the
front matter on Sheet 2 — so Sheet 2 is scraped as a regex fallback for anything Sheet 1
didn't have (see FALLBACK_PATTERNS below), each flagged with a warning so it's clear the value
was inferred from boilerplate rather than a labeled field.
"""
from __future__ import annotations

import re

import openpyxl

from app.domains.projects.po_intake import normalize as n

_LABELS = {
    "title:": "project_title",
    "author(s):": "author_names",
    "isbn:": "isbn_no",
    "e-isbn:": "ebook_isbn",
    "trim size:": "trim_size",
    "color:": "color",
    "turnover date:": "turnover_date",
    "final press pdf due:": "due_date",
}

_ISBN_RE = re.compile(r"ISBN:?\s*([\d][\d\-]{8,16}[\dXx])")
_COPYRIGHT_RE = re.compile(r"(?:©|Â©)\s*(\d{4})")


def _find_label_values(ws) -> dict[str, object]:
    """Returns raw (un-stringified) neighbor values, keyed by target field name.

    Keeping the raw value (rather than pre-cleaning to a string) matters for the two date
    fields: openpyxl hands back real `datetime` objects for date-formatted cells, and only
    `normalize.parse_date_loose` — not `normalize.clean` — knows how to handle those.
    """
    found: dict[str, object] = {}
    for row in ws.iter_rows():
        for cell in row:
            label = n.clean(cell.value)
            if not label:
                continue
            key = _LABELS.get(label.casefold())
            if not key or key in found:
                continue
            for offset in (1, 2, 3):
                neighbor = ws.cell(row=cell.row, column=cell.column + offset).value
                if isinstance(neighbor, (int, float)) or n.clean(neighbor):
                    found[key] = neighbor
                    break
    return found


def _find_totals_row(ws) -> int | None:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip().upper().startswith("TOTALS"):
                return cell.row
    return None


def _extract_chapter_table(ws) -> tuple[int | None, int | None, list[dict]]:
    totals_row = _find_totals_row(ws)
    if not totals_row:
        return None, None, []
    chapter_rows = []
    for row_idx in range(1, totals_row):
        chapter_no = ws.cell(row=row_idx, column=3).value  # column C
        if isinstance(chapter_no, (int, float)) and float(chapter_no).is_integer():
            pages = ws.cell(row=row_idx, column=7).value  # column G (rounded pages)
            chapter_rows.append({"chapter": int(chapter_no), "pages": n.to_int(pages)})
    total_pages = n.to_int(ws.cell(row=totals_row, column=7).value)
    return total_pages, (len(chapter_rows) or None), chapter_rows


def _sheet2_fallback(ws) -> dict:
    result: dict[str, object] = {}
    for row in ws.iter_rows():
        for cell in row:
            text = cell.value
            if not isinstance(text, str):
                continue
            if "isbn_no" not in result:
                m = _ISBN_RE.search(text)
                if m:
                    result["isbn_no"] = m.group(1)
            if "copyright_year" not in result:
                m = _COPYRIGHT_RE.search(text)
                if m:
                    result["copyright_year"] = int(m.group(1))
            if "author_names" not in result:
                m = re.search(r"\n\n(.*?)\n\n\s*\[.*?logo\]", text, re.DOTALL | re.IGNORECASE)
                if m:
                    names = [n.clean(line) for line in m.group(1).split("\n")]
                    names = [nm for nm in names if nm]
                    if names:
                        result["author_names"] = names
            if "publisher" not in result and "ARTECH HOUSE" in text.upper():
                result["publisher"] = "Artech House"
    return result


def parse(xlsx_path: str) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    warnings: list[str] = []

    sheet1 = wb.worksheets[0]
    labeled = _find_label_values(sheet1)
    total_pages, chapter_count, chapter_rows = _extract_chapter_table(sheet1)

    isbn_no, isbn_warning = n.normalize_isbn(labeled.get("isbn_no"))
    if isbn_warning:
        warnings.append(isbn_warning)

    due_date, due_date_warning = n.parse_date_loose(labeled.get("due_date"))
    if due_date_warning:
        warnings.append(f"FINAL PRESS PDF DUE: {due_date_warning}")

    author_raw = n.clean(labeled.get("author_names"))
    author_names = [author_raw] if author_raw else []
    copyright_year = None
    publisher = None

    if len(wb.worksheets) > 1:
        fallback = _sheet2_fallback(wb.worksheets[1])
        if not isbn_no and fallback.get("isbn_no"):
            isbn_no, isbn_warning = n.normalize_isbn(fallback["isbn_no"])
            warnings.append("ISBN was blank on the transmittal form — inferred from the front-matter copyright page, please double-check.")
        if not author_names and fallback.get("author_names"):
            author_names = fallback["author_names"]
            warnings.append("Author name(s) were blank on the transmittal form — inferred from the front-matter copyright page, please double-check.")
        copyright_year = fallback.get("copyright_year")
        publisher = fallback.get("publisher")

    turnover_date, turnover_warning = n.parse_date_loose(labeled.get("turnover_date"))
    if turnover_warning:
        warnings.append(f"Turnover Date: {turnover_warning}")

    fields = {
        "project_title": n.clean(labeled.get("project_title")),
        "isbn_no": isbn_no,
        "trim_size": n.clean(labeled.get("trim_size")),
        "color": n.clean(labeled.get("color")),
        "due_date": due_date,
        "manuscript_pages": total_pages,
        "chapter_count": chapter_count,
        "copyright_year": copyright_year,
    }

    extras = {
        "author_names": author_names,
        "ebook_isbn": n.clean(labeled.get("ebook_isbn")),
        "publisher": publisher,
        "key_dates": {
            "turnover_date": turnover_date,
        },
        "chapter_breakdown": chapter_rows,
    }

    return {"fields": fields, "extras": extras, "warnings": warnings}
