"""Figure PDF generation for the Art track.

For a given chapter, collect every figure — both images embedded in any DOCX
file in the chapter (in document order) and standalone image files in the
Art folder — and bundle them into a single PDF with one figure per page.
Each page carries a metadata table on top and the image centered below. The
resulting PDF is registered as a new ``category="Art"`` File so it appears
alongside the source figures.
"""
from __future__ import annotations

import io
import logging
import os
import re
import shutil
import subprocess
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple

try:
    import pymupdf as fitz
except ImportError:
    import fitz  # PyMuPDF
from lxml import etree as ET
from PIL import ExifTags, Image
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)

from app import models
from app.domains.projects.models import Project
from app.utils.timezone import now_ist_naive

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".tif", ".tiff", ".bmp", ".eps"}
DOCX_EXTS = {".docx", ".docm"}

# US Letter, portrait — points (1 pt = 1/72 in)
PAGE_WIDTH = 612.0
PAGE_HEIGHT = 792.0
MARGIN = 36.0
TABLE_ROW_HEIGHT = 14.0
TABLE_LABEL_WIDTH = 140.0
IMAGE_TOP_PADDING = 28.0

_WML_A = "http://schemas.openxmlformats.org/drawingml/2006/main"
_WML_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_WML_WP = "http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing"
_WML_W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
_WML_PIC = "http://schemas.openxmlformats.org/drawingml/2006/picture"
# The wpg (Word Processing Group) namespace has two flavours in the wild —
# the 2010 Microsoft one for modern grouped drawings and the OOXML one.
_WML_WPG_MS = "http://schemas.microsoft.com/office/word/2010/wordprocessingGroup"
_WML_WPG_ML = "http://schemas.openxmlformats.org/drawingml/2006/wordprocessingGroup"

# EMU per point: 914400 EMU = 1 inch = 72 pt, so 12700 EMU = 1 pt.
_EMU_PER_PT = 12700
_EMU_PER_INCH = 914400
# Resolution at which we rasterize composited groups. 300 gives print-quality
# without ballooning file size for typical figure sizes.
_COMPOSITE_DPI = 300


def _natural_key(s: str):
    return [int(t) if t.isdigit() else t.lower() for t in re.split(r"(\d+)", s or "")]


def _fmt_size(n: int) -> str:
    size = float(n)
    for unit in ("B", "KB", "MB", "GB"):
        if size < 1024 or unit == "GB":
            return f"{int(size)} B" if unit == "B" else f"{size:.1f} {unit}"
        size /= 1024.0
    return f"{size:.1f} TB"


def _fmt_ts(ts: Optional[float]) -> str:
    if ts is None:
        return "—"
    try:
        return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return "—"


class _ImageSource:
    """One figure to render — either loaded from disk or embedded in a DOCX."""

    __slots__ = ("filename", "data", "path", "source_label", "source_ts", "display_size_pt")

    def __init__(
        self,
        *,
        filename: str,
        data: Optional[bytes],
        path: Optional[str],
        source_label: str,
        source_ts: Optional[Tuple[float, float]],
        display_size_pt: Optional[Tuple[float, float]] = None,
    ):
        self.filename = filename
        self.data = data
        self.path = path
        self.source_label = source_label
        self.source_ts = source_ts
        # Author-intended display size in points (from the DOCX drawing extent).
        # None for standalone image files where no author sizing exists.
        self.display_size_pt = display_size_pt

    def open_pil(self) -> Image.Image:
        if self.data is not None:
            img = Image.open(io.BytesIO(self.data))
        else:
            img = Image.open(self.path)  # type: ignore[arg-type]

        if getattr(img, "format", None) == "EPS" or (self.filename and self.filename.lower().endswith(".eps")):
            scale = 4
            while scale > 1 and (img.width * scale) * (img.height * scale) > 100_000_000:
                scale -= 1
            try:
                img.load(scale=scale)  # type: ignore
            except Exception:
                pass
        return img

    def bytes(self) -> bytes:
        if self.data is not None:
            return self.data
        with open(self.path, "rb") as f:  # type: ignore[arg-type]
            return f.read()

    def size_bytes(self) -> int:
        if self.data is not None:
            return len(self.data)
        return os.path.getsize(self.path)  # type: ignore[arg-type]


def _drawing_extent_emu(drawing: ET._Element) -> Optional[tuple[int, int]]:
    ext = drawing.find(f".//{{{_WML_WP}}}extent")
    if ext is None:
        return None
    try:
        cx = int(ext.get("cx", "0"))
        cy = int(ext.get("cy", "0"))
    except (TypeError, ValueError):
        return None
    return (cx, cy) if cx > 0 and cy > 0 else None


def _pic_xfrm_emu(pic: ET._Element) -> Optional[tuple[int, int, int, int]]:
    """Return (offset_x, offset_y, ext_cx, ext_cy) in EMU for a <pic:pic>."""
    spPr = pic.find(f"{{{_WML_PIC}}}spPr")
    if spPr is None:
        return None
    xfrm = spPr.find(f"{{{_WML_A}}}xfrm")
    if xfrm is None:
        return None
    off = xfrm.find(f"{{{_WML_A}}}off")
    ex = xfrm.find(f"{{{_WML_A}}}ext")
    if off is None or ex is None:
        return None
    try:
        return (
            int(off.get("x", "0")),
            int(off.get("y", "0")),
            int(ex.get("cx", "0")),
            int(ex.get("cy", "0")),
        )
    except (TypeError, ValueError):
        return None


def _render_drawing_via_soffice(
    docx_path: str, drawing: ET._Element, sect_pr: Optional[ET._Element]
) -> Optional[bytes]:
    """Render a single drawing to PNG bytes by round-tripping through
    LibreOffice — for grouped drawings that mix rasters and vector shapes
    (DrawingML custGeom paths etc.) that we can't composite in Python.

    Creates a minimal DOCX containing only this drawing (preserving the
    original section properties for page size), converts to PDF via
    ``soffice --headless``, rasterizes page 1, and crops the surrounding
    whitespace. Slow (~30-40s per call for LibreOffice's startup), so callers
    should only invoke this for drawings that genuinely need vector rendering.
    """
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        return None

    try:
        with zipfile.ZipFile(docx_path, "r") as zin:
            files = {name: zin.read(name) for name in zin.namelist()}
    except (zipfile.BadZipFile, OSError):
        return None

    doc_xml_bytes = files.get("word/document.xml")
    if not doc_xml_bytes:
        return None

    try:
        doc_tree = ET.fromstring(doc_xml_bytes)
    except ET.XMLSyntaxError:
        return None

    body = doc_tree.find(f"{{{_WML_W}}}body")
    if body is None:
        return None

    # Preserve original section properties (page size / margins) so the
    # drawing renders on a page big enough to contain it; otherwise
    # LibreOffice defaults to A4 which may crop wide figures. Strip
    # header / footer references so the DOCX's running header text doesn't
    # bleed into the figure crop.
    for child in list(body):
        body.remove(child)
    p = ET.SubElement(body, f"{{{_WML_W}}}p")
    r = ET.SubElement(p, f"{{{_WML_W}}}r")
    r.append(drawing)
    if sect_pr is not None:
        for ref_tag in ("headerReference", "footerReference"):
            for ref in sect_pr.findall(f"{{{_WML_W}}}{ref_tag}"):
                sect_pr.remove(ref)
        body.append(sect_pr)

    try:
        new_doc_xml = ET.tostring(
            doc_tree, xml_declaration=True, encoding="UTF-8", standalone=True
        )
    except Exception:
        return None
    files["word/document.xml"] = new_doc_xml

    with tempfile.TemporaryDirectory() as td:
        mini_docx = os.path.join(td, "figure.docx")
        try:
            with zipfile.ZipFile(mini_docx, "w", zipfile.ZIP_DEFLATED) as zout:
                for name, data in files.items():
                    zout.writestr(name, data)
        except OSError:
            return None

        # A dedicated user-profile dir keeps concurrent conversions from
        # trampling one another's soffice lockfiles.
        profile_dir = os.path.join(td, "sofficeprofile")
        try:
            result = subprocess.run(
                [
                    soffice,
                    "--headless",
                    "--nologo",
                    "--norestore",
                    f"-env:UserInstallation=file://{profile_dir}",
                    "--convert-to",
                    "pdf",
                    "--outdir",
                    td,
                    mini_docx,
                ],
                capture_output=True,
                timeout=180,
            )
        except (subprocess.TimeoutExpired, OSError) as exc:
            logger.warning("soffice conversion failed for grouped drawing: %s", exc)
            return None

        pdf_path = os.path.join(td, "figure.pdf")
        if not os.path.exists(pdf_path):
            logger.warning(
                "soffice produced no PDF for grouped drawing (rc=%s, stderr=%s)",
                result.returncode, result.stderr[:200] if result.stderr else b"",
            )
            return None

        try:
            with fitz.open(pdf_path) as pdf:
                if len(pdf) == 0:
                    return None
                pix = pdf[0].get_pixmap(dpi=200, alpha=False)
                png_bytes = pix.tobytes("png")
        except Exception as exc:
            logger.warning("Failed to rasterize soffice PDF: %s", exc)
            return None

    # Crop the surrounding white space so the figure fills the emitted image
    # instead of sitting in a big empty page rendered at the DOCX's paper size.
    try:
        with Image.open(io.BytesIO(png_bytes)) as im:
            im = im.convert("RGB")
            # bbox() finds non-zero pixels, so invert white → black first.
            from PIL import ImageChops, ImageOps
            gray = ImageOps.invert(im.convert("L"))
            # Ignore near-white noise by thresholding.
            gray = gray.point(lambda p: 255 if p > 8 else 0)
            bbox = gray.getbbox()
            if bbox:
                pad = 20
                left = max(0, bbox[0] - pad)
                top = max(0, bbox[1] - pad)
                right = min(im.width, bbox[2] + pad)
                bottom = min(im.height, bbox[3] + pad)
                im = im.crop((left, top, right, bottom))
            buf = io.BytesIO()
            im.save(buf, format="PNG", optimize=True)
            return buf.getvalue()
    except Exception:
        return png_bytes


def _composite_group_pics(
    canvas_emu: tuple[int, int],
    pics: list[tuple[bytes, tuple[int, int, int, int]]],
) -> Optional[bytes]:
    """Rasterize a group of positioned pictures onto a single transparent PNG.

    `canvas_emu` is the group's outer extent; each `pics` entry is
    (image_bytes, (offset_x, offset_y, ext_cx, ext_cy)) in the same EMU
    coordinate system (chOff assumed to be 0 — the common Word case).
    """
    if not pics:
        return None
    cw_emu, ch_emu = canvas_emu
    if cw_emu <= 0 or ch_emu <= 0:
        return None
    px_w = max(1, int(round(cw_emu / _EMU_PER_INCH * _COMPOSITE_DPI)))
    px_h = max(1, int(round(ch_emu / _EMU_PER_INCH * _COMPOSITE_DPI)))
    canvas = Image.new("RGBA", (px_w, px_h), (255, 255, 255, 0))
    for data, (x_emu, y_emu, w_emu, h_emu) in pics:
        if w_emu <= 0 or h_emu <= 0:
            continue
        try:
            im = Image.open(io.BytesIO(data))
            if im.mode != "RGBA":
                im = im.convert("RGBA")
        except Exception:
            continue
        px = int(round(x_emu / cw_emu * px_w))
        py = int(round(y_emu / ch_emu * px_h))
        pw = max(1, int(round(w_emu / cw_emu * px_w)))
        ph = max(1, int(round(h_emu / ch_emu * px_h)))
        try:
            im = im.resize((pw, ph), Image.LANCZOS)
        except Exception:
            continue
        canvas.paste(im, (px, py), im)
    buf = io.BytesIO()
    canvas.save(buf, format="PNG", optimize=True)
    return buf.getvalue()


def _extract_docx_images(docx_path: str) -> list[_ImageSource]:
    """Return every embedded figure in `docx_path`.

    Iterates ``<w:drawing>`` elements — each represents ONE visual figure
    placement in the document. When a drawing contains multiple pictures
    inside a ``<wpg:wgp>`` group (e.g. a Word "SmartArt"-style illustration
    stitched from several bitmap parts), we composite the parts onto a
    single canvas at their author-placed positions so the reviewer sees the
    intact artwork instead of one page per fragment.

    Fallback: when the reference walk finds nothing but ``word/media/`` has
    entries — as happens with DOCX files whose figures were dropped from the
    body but left orphaned in the zip — enumerate the media directory directly
    (natural filename order), so the user still gets every figure Word knows
    about.
    """
    st = os.stat(docx_path)
    source_ts = (getattr(st, "st_birthtime", None) or st.st_ctime, st.st_mtime)
    docx_name = os.path.basename(docx_path)

    images: list[_ImageSource] = []
    emitted_media_paths: set[str] = set()

    def _load_rels(z: zipfile.ZipFile, rels_name: str) -> dict[str, str]:
        try:
            with z.open(rels_name) as f:
                tree = ET.parse(f)
        except (KeyError, ET.XMLSyntaxError):
            return {}
        rels: dict[str, str] = {}
        for rel in tree.getroot():
            if rel.tag.rsplit("}", 1)[-1] != "Relationship":
                continue
            rid = rel.get("Id")
            target = rel.get("Target")
            if rid and target:
                rels[rid] = target
        return rels

    def _resolve_media(target: str, part_dir: str, names: set[str]) -> Optional[str]:
        # Relationship targets are relative to the part's folder; "../media/x"
        # from word/document.xml resolves to word/media/x, etc.
        raw = target.lstrip("/")
        candidate = os.path.normpath(f"{part_dir}/{raw}").replace(os.sep, "/")
        if candidate in names:
            return candidate
        alt = f"word/{raw}"
        return alt if alt in names else None

    def _blip_in_extlst(blip: ET._Element) -> bool:
        # Skip fallback/alt blips nested inside <a:extLst> (e.g. SVG extensions)
        # — they're alternate encodings of a parent blip, not separate figures.
        anc = blip.getparent()
        while anc is not None:
            if anc.tag == f"{{{_WML_A}}}extLst":
                return True
            anc = anc.getparent()
        return False

    try:
        with zipfile.ZipFile(docx_path) as z:
            names = set(z.namelist())

            # Every "story" part that can hold drawings and its matching rels.
            parts: list[tuple[str, str]] = []
            for part in ("word/document.xml", "word/footnotes.xml", "word/endnotes.xml"):
                if part in names:
                    parts.append((part, f"word/_rels/{os.path.basename(part)}.rels"))
            for name in sorted(names):
                if (name.startswith("word/header") or name.startswith("word/footer")) and name.endswith(".xml"):
                    parts.append((name, f"word/_rels/{os.path.basename(name)}.rels"))

            for part_name, rels_name in parts:
                rels_map = _load_rels(z, rels_name)
                if not rels_map:
                    continue
                try:
                    with z.open(part_name) as f:
                        tree = ET.parse(f)
                except (KeyError, ET.XMLSyntaxError):
                    continue

                # Grab the section properties so we can preserve page size /
                # margins when we spawn a minimal DOCX for LibreOffice rendering.
                sect_pr_original = None
                body_el = tree.find(f"{{{_WML_W}}}body")
                if body_el is not None:
                    sp = body_el.find(f"{{{_WML_W}}}sectPr")
                    if sp is not None:
                        sect_pr_original = sp

                part_dir = os.path.dirname(part_name)
                for drawing in tree.iter(f"{{{_WML_W}}}drawing"):
                    extent_emu = _drawing_extent_emu(drawing)
                    display_size_pt = (
                        (extent_emu[0] / _EMU_PER_PT, extent_emu[1] / _EMU_PER_PT)
                        if extent_emu else None
                    )

                    # Collect every primary blip in this drawing (skipping SVG
                    # fallbacks nested in extLst). Preserve the containing
                    # <pic:pic> so we can read its per-child xfrm when we need
                    # to composite a group.
                    entries: list[tuple[str, bytes, Optional[ET._Element]]] = []
                    for blip in drawing.iter(f"{{{_WML_A}}}blip"):
                        if _blip_in_extlst(blip):
                            continue
                        rId = blip.get(f"{{{_WML_R}}}embed") or blip.get(f"{{{_WML_R}}}link")
                        if not rId:
                            continue
                        target = rels_map.get(rId)
                        if not target:
                            continue
                        media_path = _resolve_media(target, part_dir, names)
                        if not media_path:
                            continue
                        try:
                            with z.open(media_path) as mf:
                                data = mf.read()
                        except KeyError:
                            continue
                        # Find the containing <pic:pic> for per-child xfrm.
                        pic_anc = blip.getparent()
                        while pic_anc is not None and pic_anc.tag != f"{{{_WML_PIC}}}pic":
                            pic_anc = pic_anc.getparent()
                        entries.append((media_path, data, pic_anc))
                        emitted_media_paths.add(media_path)

                    if not entries:
                        continue

                    if len(entries) == 1:
                        media_path, data, _ = entries[0]
                        images.append(_ImageSource(
                            filename=os.path.basename(media_path),
                            data=data,
                            path=None,
                            source_label=docx_name,
                            source_ts=source_ts,
                            display_size_pt=display_size_pt,
                        ))
                        continue

                    # Multiple blips in one drawing → grouped artwork. If the
                    # drawing also carries vector shapes (wps:wsp with custGeom
                    # paths — arms, legs, outlines etc.), plain raster
                    # compositing loses them. Round-trip through LibreOffice
                    # in that case so the vectors render alongside the rasters.
                    has_vector_shapes = drawing.find(f".//{{{_WML_A}}}custGeom") is not None
                    rendered: Optional[bytes] = None
                    if has_vector_shapes:
                        # ET's element already lives inside the current tree —
                        # deepcopy so appending it to a fresh document doesn't
                        # detach it and break subsequent iteration.
                        import copy
                        rendered = _render_drawing_via_soffice(
                            docx_path, copy.deepcopy(drawing),
                            copy.deepcopy(sect_pr_original) if sect_pr_original is not None else None,
                        )

                    if rendered is None:
                        pics_positioned: list[tuple[bytes, tuple[int, int, int, int]]] = []
                        for _mp, data, pic_el in entries:
                            xfrm = _pic_xfrm_emu(pic_el) if pic_el is not None else None
                            if xfrm is None:
                                # No per-child xfrm — fall back to filling the canvas.
                                xfrm = (0, 0, extent_emu[0] if extent_emu else 0,
                                        extent_emu[1] if extent_emu else 0)
                            pics_positioned.append((data, xfrm))

                        if extent_emu is None:
                            # No group extent to draw into — degrade gracefully by
                            # emitting each part separately, as we would have before.
                            for media_path, data, _ in entries:
                                images.append(_ImageSource(
                                    filename=os.path.basename(media_path),
                                    data=data,
                                    path=None,
                                    source_label=docx_name,
                                    source_ts=source_ts,
                                    display_size_pt=None,
                                ))
                            continue

                        rendered = _composite_group_pics(extent_emu, pics_positioned)
                        if rendered is None:
                            continue

                    first_stem = Path(entries[0][0]).stem
                    images.append(_ImageSource(
                        filename=f"{first_stem}_group.png",
                        data=rendered,
                        path=None,
                        source_label=docx_name,
                        source_ts=source_ts,
                        display_size_pt=display_size_pt,
                    ))

            # Fallback: any media entry not already emitted via a reference —
            # covers DOCX files whose figures are orphaned in the zip. Sorted
            # naturally so image1, image2, …, image10 stay in order.
            leftovers = sorted(
                (n for n in names if n.startswith("word/media/") and not n.endswith("/") and n not in emitted_media_paths),
                key=lambda n: _natural_key(os.path.basename(n)),
            )
            for media_path in leftovers:
                try:
                    with z.open(media_path) as mf:
                        data = mf.read()
                except KeyError:
                    continue
                images.append(_ImageSource(
                    filename=os.path.basename(media_path),
                    data=data,
                    path=None,
                    source_label=docx_name,
                    source_ts=source_ts,
                ))
    except zipfile.BadZipFile:
        return []

    return images


_COLOR_MODE_LABELS = {
    "1": "Bilevel (1-bit B/W)",
    "L": "Grayscale (8-bit)",
    "LA": "Grayscale with alpha",
    "P": "Palette (indexed)",
    "PA": "Palette with alpha",
    "RGB": "RGB",
    "RGBA": "RGB with alpha",
    "CMYK": "CMYK",
    "YCbCr": "YCbCr",
    "LAB": "L*a*b*",
    "HSV": "HSV",
    "I": "Grayscale (32-bit int)",
    "F": "Grayscale (32-bit float)",
    "I;16": "Grayscale (16-bit)",
}


def _describe_color_space(im) -> Optional[str]:
    label = _COLOR_MODE_LABELS.get(im.mode, im.mode)
    profile_name = None
    icc = im.info.get("icc_profile")
    if icc:
        # Parse the "desc" tag out of the ICC profile if PIL didn't already
        # expose a description. Cheap fallback: just note that a profile exists.
        try:
            from PIL import ImageCms
            prof = ImageCms.ImageCmsProfile(io.BytesIO(icc))
            profile_name = ImageCms.getProfileDescription(prof).strip() or None
        except Exception:
            profile_name = "embedded ICC profile"
    if profile_name:
        return f"{label} · {profile_name}"
    return label


def _extract_metadata(src: _ImageSource) -> list[tuple[str, str]]:
    width = height = None
    dpi_x = dpi_y = None
    fmt = color_space = exif_dt = None
    try:
        with src.open_pil() as im:
            width, height = im.size
            fmt = im.format
            color_space = _describe_color_space(im)
            dpi_val = im.info.get("dpi") or im.info.get("jfif_density")
            if dpi_val and isinstance(dpi_val, (tuple, list)) and len(dpi_val) >= 2:
                try:
                    dx, dy = float(dpi_val[0]), float(dpi_val[1])
                    if dx > 0:
                        dpi_x = int(round(dx))
                    if dy > 0:
                        dpi_y = int(round(dy))
                except (TypeError, ValueError):
                    pass
            try:
                exif = im.getexif()
                if exif:
                    for tag_id, val in exif.items():
                        tag = ExifTags.TAGS.get(tag_id, tag_id)
                        if tag in ("DateTimeOriginal", "DateTime") and isinstance(val, str):
                            exif_dt = val
                            break
                    if dpi_x is None:
                        # EXIF stores resolution in tags 282 (XResolution) and
                        # 283 (YResolution). Unit tag 296: 2=inch, 3=cm.
                        xres = exif.get(282)
                        yres = exif.get(283)
                        unit = exif.get(296, 2)
                        def _to_dpi(v):
                            try:
                                v = float(v)
                            except (TypeError, ValueError):
                                return None
                            if unit == 3:  # per-cm → per-inch
                                v *= 2.54
                            return int(round(v)) if v > 0 else None
                        dpi_x = _to_dpi(xres)
                        dpi_y = _to_dpi(yres)
            except Exception:
                pass
    except Exception:
        pass

    # Physical dimensions in picas (1 pica = 1/6 inch, so pixels/DPI * 6).
    # Falls back to 96 DPI (Word's default screen DPI) when the image doesn't
    # declare one — otherwise we'd have to hide dimensions for every plain PNG.
    if width and height:
        eff_dpi = dpi_x or dpi_y or 96
        w_pc = width / eff_dpi * 6.0
        h_pc = height / eff_dpi * 6.0
        dimensions = f"{w_pc:.1f} × {h_pc:.1f} pc"
    else:
        dimensions = "N/A"

    if dpi_x and dpi_y:
        dpi_str = f"{dpi_x} DPI" if dpi_x == dpi_y else f"{dpi_x} × {dpi_y} DPI"
    elif dpi_x or dpi_y:
        dpi_str = f"{dpi_x or dpi_y} DPI"
    else:
        dpi_str = "N/A"

    fmt_str = fmt or Path(src.filename).suffix.lstrip(".").upper() or "N/A"

    ctime = mtime = None
    if src.source_ts:
        ctime, mtime = src.source_ts

    def _fmt_ts_or_na(ts):
        s = _fmt_ts(ts)
        return s if s != "—" else "N/A"

    return [
        ("File Name", src.filename or "N/A"),
        ("Image Format", fmt_str),
        ("Dimensions", dimensions),
        ("Resolution (DPI)", dpi_str),
        ("Color Space", color_space or "N/A"),
        ("File Size", _fmt_size(src.size_bytes())),
        ("Creation Date", exif_dt or _fmt_ts_or_na(ctime)),
        ("Last Modified Date", _fmt_ts_or_na(mtime)),
    ]


def _draw_metadata_table(page: fitz.Page, rows: list[tuple[str, str]], top: float) -> float:
    left = MARGIN
    right = PAGE_WIDTH - MARGIN
    label_x = left + TABLE_LABEL_WIDTH
    bottom = top + TABLE_ROW_HEIGHT * len(rows)

    page.draw_rect(fitz.Rect(left, top, right, bottom), color=(0, 0, 0), width=0.6)
    page.draw_line(fitz.Point(label_x, top), fitz.Point(label_x, bottom), color=(0, 0, 0), width=0.4)

    for i, (label, value) in enumerate(rows):
        row_top = top + i * TABLE_ROW_HEIGHT
        if i > 0:
            page.draw_line(
                fitz.Point(left, row_top),
                fitz.Point(right, row_top),
                color=(0, 0, 0), width=0.25,
            )
        text_y = row_top + TABLE_ROW_HEIGHT - 4
        page.insert_text(fitz.Point(left + 4, text_y), label, fontsize=8, fontname="hebo")
        display = str(value)
        if len(display) > 140:
            display = display[:137] + "…"
        page.insert_text(fitz.Point(label_x + 4, text_y), display, fontsize=8, fontname="helv")

    return bottom


def _fit_rect_in_area(width_pt: float, height_pt: float, area: fitz.Rect) -> fitz.Rect:
    """Center a ``width_pt × height_pt`` box inside ``area``, top-aligned."""
    area_w = area.x1 - area.x0
    area_h = area.y1 - area.y0
    if width_pt <= 0 or height_pt <= 0:
        return fitz.Rect(area.x0, area.y0, area.x1, area.y1)
    scale = min(area_w / width_pt, area_h / height_pt, 1.0)
    tw = width_pt * scale
    th = height_pt * scale
    cx = (area.x0 + area.x1) / 2
    return fitz.Rect(cx - tw / 2, area.y0, cx + tw / 2, area.y0 + th)


def _insert_eps_vector(page: fitz.Page, src: _ImageSource, area: fitz.Rect) -> bool:
    """Embed EPS as vector by round-tripping through Ghostscript's pdfwrite."""
    from app.domains.files import eps_service

    try:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp_pdf = Path(tmp.name)
        try:
            eps_service.eps_to_pdf(Path(src.path), tmp_pdf)  # type: ignore[arg-type]
            with fitz.open(str(tmp_pdf)) as ep:
                if ep.page_count == 0:
                    return False
                first = ep.load_page(0)
                pw, ph = first.rect.width, first.rect.height
                target = _fit_rect_in_area(pw, ph, area)
                page.show_pdf_page(target, ep, 0)
            return True
        finally:
            try:
                tmp_pdf.unlink(missing_ok=True)  # type: ignore[call-arg]
            except TypeError:
                if tmp_pdf.exists():
                    tmp_pdf.unlink()
    except Exception as exc:
        logger.warning("EPS vector embed failed for %s: %s", src.filename, exc)
        return False


def _insert_eps_raster(page: fitz.Page, src: _ImageSource, area: fitz.Rect) -> bool:
    """Fallback: rasterize EPS via Ghostscript at a target-sized DPI."""
    from app.domains.files import eps_service

    try:
        bbox = eps_service.read_bbox_points(Path(src.path))  # type: ignore[arg-type]
        area_w = area.x1 - area.x0
        area_h = area.y1 - area.y0
        if bbox:
            wpt, hpt = bbox
            scale = min(area_w / wpt, area_h / hpt, 1.0)
            target_pt_edge = max(wpt, hpt) * scale
        else:
            target_pt_edge = max(area_w, area_h)
        # 300 DPI at the drawn size — print-quality without runaway pixel counts.
        target_px_edge = max(600, int(round((target_pt_edge / 72.0) * 300)))
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            tmp_png = Path(tmp.name)
        try:
            eps_service.eps_to_png(
                Path(src.path),  # type: ignore[arg-type]
                tmp_png,
                target_px_edge=target_px_edge,
            )
            with Image.open(tmp_png) as im:
                w, h = im.size
                buf = io.BytesIO()
                if im.mode not in ("RGB", "RGBA", "L"):
                    im = im.convert("RGBA")
                im.save(buf, format="PNG")
            target = _fit_rect_in_area(w, h, area)
            page.insert_image(target, stream=buf.getvalue())
            return True
        finally:
            try:
                tmp_png.unlink(missing_ok=True)  # type: ignore[call-arg]
            except TypeError:
                if tmp_png.exists():
                    tmp_png.unlink()
    except Exception as exc:
        logger.warning("EPS high-DPI raster failed for %s: %s", src.filename, exc)
        return False


def _insert_image_centered(page: fitz.Page, src: _ImageSource, area: fitz.Rect) -> None:
    # Standalone EPS on disk gets a vector-preserving path: render to PDF via
    # Ghostscript and embed via show_pdf_page, so the figure stays resolution-
    # independent. Fall back to high-DPI Ghostscript raster if vector embed
    # fails, and only then to the Pillow path below.
    ext = Path(src.filename or "").suffix.lower() if src.filename else ""
    if ext == ".eps" and src.path and src.data is None:
        if _insert_eps_vector(page, src, area):
            return
        if _insert_eps_raster(page, src, area):
            return

    w = h = None
    try:
        with src.open_pil() as im:
            w, h = im.size
    except Exception:
        pass

    area_w = area.x1 - area.x0
    area_h = area.y1 - area.y0

    # Prefer the author's intended display size from the DOCX drawing extent.
    # This matches what Word shows and avoids stretching a small icon to fill
    # the page (which either pixelates or blurs it). Fall back to fit-to-page
    # for standalone image files where no author sizing exists.
    if src.display_size_pt and src.display_size_pt[0] > 0 and src.display_size_pt[1] > 0:
        tw_target, th_target = src.display_size_pt
        # Clamp to the available area (never overflow the page).
        clamp = min(area_w / tw_target, area_h / th_target, 1.0)
        tw, th = tw_target * clamp, th_target * clamp
        target_pixel_scale = (tw / w) if (w and w > 0) else 1.0
    elif w and h and w > 0 and h > 0:
        fit_scale = min(area_w / w, area_h / h)
        tw, th = w * fit_scale, h * fit_scale
        target_pixel_scale = fit_scale
    else:
        tw, th = area_w, area_h
        target_pixel_scale = 1.0

    # Horizontally centered on the page; top-aligned within the area so the
    # image sits directly below the metadata table instead of drifting to the
    # vertical middle of a mostly-empty half-page.
    cx = (area.x0 + area.x1) / 2
    top_y = area.y0
    target = fitz.Rect(cx - tw / 2, top_y, cx + tw / 2, top_y + th)

    data = src.bytes()
    # If the target size requires enlarging the source bitmap, pre-upscale via
    # Pillow LANCZOS so PDF viewers don't render it with nearest-neighbor.
    if target_pixel_scale > 1.0 and w and h:
        try:
            with src.open_pil() as im:
                if im.mode not in ("RGB", "RGBA", "L"):
                    im = im.convert("RGBA")
                new_w = max(1, int(round(w * target_pixel_scale)))
                new_h = max(1, int(round(h * target_pixel_scale)))
                im = im.resize((new_w, new_h), Image.LANCZOS)
                buf = io.BytesIO()
                im.save(buf, format="PNG")
                data = buf.getvalue()
        except Exception:
            pass

    try:
        page.insert_image(target, stream=data)
        return
    except Exception:
        pass

    # Fallback: PyMuPDF struggles with certain TIFFs/EPS — re-encode via Pillow.
    try:
        with src.open_pil() as im:
            if im.mode not in ("RGB", "RGBA", "L"):
                im = im.convert("RGB")
            buf = io.BytesIO()
            im.save(buf, format="PNG")
            page.insert_image(target, stream=buf.getvalue())
    except Exception:
        pass


def _collect_images_from_file(source: models.File) -> list[_ImageSource]:
    if not source.path or not os.path.exists(source.path):
        return []
    ext = Path(source.filename or source.path).suffix.lower()
    if ext in DOCX_EXTS:
        return _extract_docx_images(source.path)
    if ext in IMAGE_EXTS:
        try:
            st = os.stat(source.path)
            ts = (getattr(st, "st_birthtime", None) or st.st_ctime, st.st_mtime)
        except OSError:
            ts = None
        return [_ImageSource(
            filename=source.filename or os.path.basename(source.path),
            data=None,
            path=source.path,
            source_label=source.filename or os.path.basename(source.path),
            source_ts=ts,
        )]
    return []


def _collect_images_from_chapter(db: Session, chapter_id: int) -> list[_ImageSource]:
    all_files = (
        db.query(models.File)
        .filter(models.File.chapter_id == chapter_id)
        .all()
    )

    sources: list[_ImageSource] = []

    # 1. Every DOCX in the chapter (any category) — contributes its embedded
    #    images in document order.
    docx_files = [
        f for f in all_files
        if f.path
        and os.path.exists(f.path)
        and Path(f.filename or f.path).suffix.lower() in DOCX_EXTS
    ]
    docx_files.sort(key=lambda f: _natural_key(f.filename or ""))
    for f in docx_files:
        sources.extend(_extract_docx_images(f.path))

    # 2. Standalone image files already in the Art folder — appended after
    #    DOCX-extracted figures.
    art_images = [
        f for f in all_files
        if f.category == "Art"
        and f.path
        and os.path.exists(f.path)
        and Path(f.filename or f.path).suffix.lower() in IMAGE_EXTS
    ]
    art_images.sort(key=lambda f: _natural_key(f.filename or ""))
    for f in art_images:
        try:
            st = os.stat(f.path)
            ts = (getattr(st, "st_birthtime", None) or st.st_ctime, st.st_mtime)
        except OSError:
            ts = None
        sources.append(_ImageSource(
            filename=f.filename or os.path.basename(f.path),
            data=None,
            path=f.path,
            source_label="Art folder",
            source_ts=ts,
        ))

    return sources


def generate_figure_pdf(
    db: Session,
    *,
    project_id: int,
    chapter_id: int,
    actor_user_id: Optional[int],
    upload_dir: str,
    source_file_id: Optional[int] = None,
) -> Tuple[Optional[models.File], int, Optional[str]]:
    """Generate a Figure PDF for the chapter.

    When ``source_file_id`` is provided, extract figures from that single
    chapter file and name the output after its stem so the Art track shows
    the source at a glance. Otherwise bundle every DOCX-embedded figure and
    standalone Art image in the chapter under a generic chapter-based name.

    Returns ``(file_record, figures_included, error_message)``.
    """
    project = db.query(Project).filter(Project.id == project_id).first()
    chapter = db.query(models.ChapterInfo).filter(models.ChapterInfo.id == chapter_id).first()
    if not project or not chapter:
        return None, 0, "Project or chapter not found."

    source_file: Optional[models.File] = None
    if source_file_id is not None:
        source_file = db.query(models.File).filter(
            models.File.id == source_file_id,
            models.File.chapter_id == chapter_id,
        ).first()
        if not source_file:
            return None, 0, "Source file not found in this chapter."
        sources = _collect_images_from_file(source_file)
        if not sources:
            return None, 0, (
                f"No figures found in {source_file.filename}. The file must be a DOCX "
                "with embedded images or an image file."
            )
    else:
        sources = _collect_images_from_chapter(db, chapter_id)
        if not sources:
            return None, 0, (
                "No figures found. Upload images to the Art folder or add a DOCX "
                "containing embedded figures to this chapter."
            )

    doc = fitz.open()
    try:
        for src in sources:
            page = doc.new_page(width=PAGE_WIDTH, height=PAGE_HEIGHT)
            rows = _extract_metadata(src)
            table_bottom = _draw_metadata_table(page, rows, MARGIN)
            image_area = fitz.Rect(
                MARGIN,
                table_bottom + IMAGE_TOP_PADDING,
                PAGE_WIDTH - MARGIN,
                PAGE_HEIGHT - MARGIN,
            )
            _insert_image_centered(page, src, image_area)

        base_path = f"{upload_dir}/{project.code}/{chapter.chapters}/Art"
        os.makedirs(base_path, exist_ok=True)

        if source_file is not None:
            source_stem = Path(source_file.filename or "source").stem or "source"
            filename = f"{source_stem}_ArtProof.pdf"
        else:
            filename = f"{chapter.chapters}_ArtProof.pdf"
        out_path = os.path.join(base_path, filename)
        doc.save(out_path, garbage=4, deflate=True)
    finally:
        doc.close()

    db_file = models.File(
        project_id=project_id,
        chapter_id=chapter_id,
        filename=filename,
        file_type="pdf",
        category="Art",
        path=out_path,
        version=1,
        is_original=False,
        uploaded_by_id=actor_user_id,
        source_file_id=source_file.id if source_file is not None else None,
    )
    db.add(db_file)
    db.commit()
    db.refresh(db_file)
    return db_file, len(sources), None


# Columns and layout follow the sample "Information Management" xlsx the team
# uses today. Plain formatting, no bold/fill/borders, empty strings (not None)
# for cells the tool cannot auto-fill so the sheet looks byte-identical to a
# freshly-created blank template.
_ASSESSMENT_COLUMNS = (
    "ART NAME",
    "FIGURE NO.",
    "PART NO.",
    "TYPE",
    "ART CLASSIFICATION",
    "SUPPLIED IMAGE COLOR",
    "SUPPLIED IMAGE MODE",
    "SUPPLIED IMAGE TYPE",
    "SUPPLIED SIZE",
    "FINAL SIZE",
    "SPECIAL INSTRUCTION ",  # trailing space matches the sample header
    "FINAL ART COLOR",
    "COMPLEXITY",
    "RESOLUTION (HALFTONE ONLY)",
    "STATUS",
)


def _fmt_size_assessment(n: int) -> str:
    """File-size formatter matching the sample ("1 MB", "236 K", "902 K")."""
    if n < 1024:
        return f"{n} B"
    if n < 1024 * 1024:
        return f"{round(n / 1024)} K"
    if n < 1024 * 1024 * 1024:
        return f"{round(n / (1024 * 1024))} MB"
    return f"{round(n / (1024 * 1024 * 1024))} GB"


_FIGURE_NUM_RE = re.compile(r"[Ff]ig(?:ure)?[\s_\-]*(\d+(?:\.\d+)*)")
_IMAGE_INDEX_RE = re.compile(r"image[\s_\-]*0*(\d+)", re.IGNORECASE)


def _extract_figure_number(filename: str, chapter_prefix: str) -> str:
    """Best-effort figure number for the FIGURE NO. column.

    Matches ``Figure3.1.jpg`` → ``3.1`` directly, or combines the chapter
    prefix with an ``imageNNN`` index (``ch003-image001.tif`` with prefix
    ``3`` → ``3.1``). Returns "" when nothing sensible can be extracted.
    """
    stem = Path(filename or "").stem
    m = _FIGURE_NUM_RE.search(stem)
    if m:
        return m.group(1)
    m = _IMAGE_INDEX_RE.search(stem)
    if m and chapter_prefix:
        return f"{chapter_prefix}.{int(m.group(1))}"
    return ""


_COLOR_SPACE_ONLY = {
    "1": "B&W",
    "L": "Grayscale",
    "LA": "Grayscale",
    "P": "Indexed",
    "PA": "Indexed",
    "RGB": "RGB",
    "RGBA": "RGB",
    "CMYK": "CMYK",
    "YCbCr": "RGB",
    "I": "Grayscale",
    "F": "Grayscale",
}


def _assessment_row_values(src: "_ImageSource", chapter_prefix: str) -> list[str]:
    """Build the 15-column row for an image source, in _ASSESSMENT_COLUMNS order."""
    width = height = None
    dpi_x = dpi_y = None
    pil_mode = fmt = color_space = None
    size_bytes = src.size_bytes()
    try:
        with src.open_pil() as im:
            width, height = im.size
            fmt = im.format
            pil_mode = im.mode
            color_space = _COLOR_SPACE_ONLY.get(im.mode, im.mode)
            dpi_val = im.info.get("dpi") or im.info.get("jfif_density")
            if dpi_val and isinstance(dpi_val, (tuple, list)) and len(dpi_val) >= 2:
                try:
                    dx, dy = float(dpi_val[0]), float(dpi_val[1])
                    if dx > 0:
                        dpi_x = int(round(dx))
                    if dy > 0:
                        dpi_y = int(round(dy))
                except (TypeError, ValueError):
                    pass
    except Exception:
        pass

    fmt_str = (fmt or Path(src.filename).suffix.lstrip(".").upper()) if src.filename else ""

    if dpi_x and dpi_y and dpi_x == dpi_y:
        dpi_str = f"{dpi_x} DPI"
    elif dpi_x and dpi_y:
        dpi_str = f"{dpi_x} × {dpi_y} DPI"
    elif dpi_x or dpi_y:
        dpi_str = f"{dpi_x or dpi_y} DPI"
    else:
        dpi_str = ""

    figure_no = _extract_figure_number(src.filename or "", chapter_prefix)
    is_image = Path(src.filename or "").suffix.lower() in IMAGE_EXTS

    return [
        src.filename or "",                               # ART NAME
        figure_no,                                        # FIGURE NO.
        "",                                               # PART NO.
        "Figure" if is_image else "",                     # TYPE
        "",                                               # ART CLASSIFICATION
        color_space or "",                                # SUPPLIED IMAGE COLOR
        pil_mode or "",                                   # SUPPLIED IMAGE MODE
        (fmt_str or "").upper(),                          # SUPPLIED IMAGE TYPE
        _fmt_size_assessment(size_bytes),                 # SUPPLIED SIZE
        "",                                               # FINAL SIZE
        "",                                               # SPECIAL INSTRUCTION
        "",                                               # FINAL ART COLOR
        "",                                               # COMPLEXITY
        dpi_str,                                          # RESOLUTION (HALFTONE ONLY)
        "",                                               # STATUS
    ]


def generate_figure_assessment(
    db: Session,
    *,
    project_id: int,
    chapter_id: int,
    actor_user_id: Optional[int],
    upload_dir: str,
    source_file_id: Optional[int] = None,
) -> Tuple[Optional[models.File], int, Optional[str]]:
    """Generate a Figure Assessment xlsx for the chapter.

    Layout matches the team's "Information Management" template — 15 columns,
    plain Calibri 12 with no header styling, no frozen panes, empty strings for
    blank cells. Auto-fills what can be derived from image metadata (ART NAME,
    FIGURE NO., TYPE, SUPPLIED IMAGE COLOR/MODE/TYPE, SUPPLIED SIZE, RESOLUTION
    for images with a real DPI) and leaves the manually-managed columns blank.
    """
    from openpyxl import Workbook

    project = db.query(Project).filter(Project.id == project_id).first()
    chapter = db.query(models.ChapterInfo).filter(models.ChapterInfo.id == chapter_id).first()
    if not project or not chapter:
        return None, 0, "Project or chapter not found."

    source_file: Optional[models.File] = None
    if source_file_id is not None:
        source_file = db.query(models.File).filter(
            models.File.id == source_file_id,
            models.File.chapter_id == chapter_id,
        ).first()
        if not source_file:
            return None, 0, "Source file not found in this chapter."
        sources = _collect_images_from_file(source_file)
        if not sources:
            return None, 0, (
                f"No figures found in {source_file.filename}. The file must be a DOCX "
                "with embedded images or an image file."
            )
    else:
        sources = _collect_images_from_chapter(db, chapter_id)
        if not sources:
            return None, 0, (
                "No figures found. Upload images to the Art folder or add a DOCX "
                "containing embedded figures to this chapter."
            )

    chapter_num_match = re.search(r"(\d+)", chapter.chapters or "")
    chapter_prefix = str(int(chapter_num_match.group(1))) if chapter_num_match else ""

    sheet_title_raw = (
        f"{chapter_prefix}. {chapter.chapter_title}"
        if chapter_prefix and chapter.chapter_title
        else (chapter.chapter_title or chapter.chapters or "Figure Assessment")
    )
    # Excel sheet titles: max 31 chars, no []:*?/\
    sheet_title = re.sub(r"[\[\]:*?/\\]", "", sheet_title_raw)[:31]

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    for col_idx, header in enumerate(_ASSESSMENT_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, src in enumerate(sources, start=2):
        for col_idx, value in enumerate(_assessment_row_values(src, chapter_prefix), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    base_path = f"{upload_dir}/{project.code}/{chapter.chapters}/Art"
    os.makedirs(base_path, exist_ok=True)

    if source_file is not None:
        source_stem = Path(source_file.filename or "source").stem or "source"
        filename = f"{source_stem}_ArtAssessment.xlsx"
    else:
        filename = f"{chapter.chapters}_ArtAssessment.xlsx"
    out_path = os.path.join(base_path, filename)
    wb.save(out_path)

    db_file = models.File(
        project_id=project_id,
        chapter_id=chapter_id,
        filename=filename,
        file_type="xlsx",
        category="Art",
        path=out_path,
        version=1,
        is_original=False,
        uploaded_by_id=actor_user_id,
        source_file_id=source_file.id if source_file is not None else None,
    )
    db.add(db_file)
    db.commit()
    db.refresh(db_file)
    return db_file, len(sources), None
