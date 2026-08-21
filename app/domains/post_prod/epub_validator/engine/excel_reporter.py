import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

def generate_gwp_report(
    validation_result: dict, 
    template_path: str, 
    output_xlsx_path: str, 
    assignee_name: str = "", 
    uploaded_date: str = "",
    epubcheck_result: dict = None,
    ace_result: dict = None
):
    """Generates an Excel report matching the GWP template from the validation result."""
    
    with open(template_path, "r") as f:
        template = json.load(f)
        
    wb = Workbook()
    ws = wb.active
    ws.title = "QA Checklist"
    
    # 1. Title
    ws.append([template.get("title", "")])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    # 2. Metadata row (e.g. Date file rec'd...)
    metadata_row = template.get("metadata_row", [])
    current_date = datetime.now().strftime('%m/%d/%y')
    
    updated_metadata = []
    for cell in metadata_row:
        if "{date}" in cell:
            cell = cell.replace("{date}", current_date)
        if "{uploaded_date}" in cell:
            # Replace with the uploaded date; fallback to current date if missing
            display_date = uploaded_date if uploaded_date else current_date
            cell = cell.replace("{uploaded_date}", display_date)
        if "{assignee name}" in cell:
            display_name = assignee_name if assignee_name else "Unknown"
            cell = cell.replace("{assignee name}", display_name)
        updated_metadata.append(cell)
            
    ws.append(updated_metadata)
    
    # 3. Headers
    # The actual headers are usually the first criteria in the 'Uncategorized' category.
    # But let's check if there's a headers array
    
    current_row = 3
    
    # Map rule_id -> issues from validation_result
    # The result has "files": [{ "rule_id": "...", "result": {"issues_count": N, "issues": [...] } }]
    rule_issues = {}
    for f in validation_result.get("files", []):
        rid = f.get("rule_id")
        res = f.get("result", {})
        issues = res.get("issues", [])
        if rid:
            if rid not in rule_issues:
                rule_issues[rid] = []
            
            # Format the issues to include the file path
            file_name = f.get("file_details", {}).get("file_name", "")
            for i in issues:
                msg = i.get("message", "")
                if file_name and file_name != "[book-level]":
                    rule_issues[rid].append(f"[{file_name}] {msg}")
                else:
                    rule_issues[rid].append(msg)
                    
    # Helper to evaluate status based on mapped rules
    def evaluate_criteria(criteria):
        rule_ids = criteria.get("rule_ids", [])
        if not isinstance(rule_ids, list):
            rule_ids = [rule_ids]
            
        single_rule = criteria.get("rule_id")
        if single_rule and single_rule not in rule_ids:
            rule_ids.append(single_rule)
            
        # Virtual rule checks (works for any customer template that maps these IDs)
        if "EPUBCHECK" in rule_ids and epubcheck_result:
            return epubcheck_result.get("status", "Yet to check"), epubcheck_result.get("notes", "")
        if "ACE_CHECK" in rule_ids and ace_result:
            return ace_result.get("status", "Yet to check"), ace_result.get("notes", "")
            
        if not rule_ids:
            return criteria.get("default_status", ""), criteria.get("default_notes", "")
            
        all_issues = []
        rules_executed = False
        
        for rid in rule_ids:
            if rid in rule_issues:
                rules_executed = True
                all_issues.extend(rule_issues[rid])
                
        if not rules_executed:
            # Rules mapped but not found in validation results
            return criteria.get("default_status", ""), criteria.get("default_notes", "")
            
        if all_issues:
            # Issues found, mark as Fail
            notes = "\n\n".join(all_issues)
            return "Fail", notes
        else:
            return "Pass", ""

    # Build the rows
    for category in template.get("structure", []):
        cat_name = category.get("category_name")
        if cat_name != "Uncategorized":
            # Write Category header
            ws.append([cat_name])
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            ws.cell(row=current_row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            current_row += 1
            
        for criteria in category.get("criteria", []):
            name = criteria.get("name", "")
            resources = criteria.get("resources", "")
            
            # If it's the actual header row (Criteria, Pass/Fail, Notes)
            if name == "Criteria":
                ws.append([name, criteria.get("default_status", ""), criteria.get("default_notes", ""), resources])
                for col in range(1, 5):
                    ws.cell(row=current_row, column=col).font = Font(bold=True)
            else:
                status, notes = evaluate_criteria(criteria)
                ws.append([name, status, notes, resources])
                
                # Apply colors based on status
                status_cell = ws.cell(row=current_row, column=2)
                lower_status = status.lower()
                if lower_status == "pass":
                    status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    status_cell.font = Font(color="006100")
                elif lower_status == "fail":
                    status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    status_cell.font = Font(color="9C0006")
                elif lower_status == "yet to check":
                    status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    status_cell.font = Font(color="9C5700")
                elif lower_status == "n/a":
                    status_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                
            # Alignment for all cells in this row
            for col in range(1, 5):
                ws.cell(row=current_row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
                
            current_row += 1
            
    # Set column widths
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 50
    
    ws.sheet_view.showGridLines = True
    
    # Make sure output directory exists
    os.makedirs(os.path.dirname(output_xlsx_path), exist_ok=True)
    wb.save(output_xlsx_path)
    return output_xlsx_path
