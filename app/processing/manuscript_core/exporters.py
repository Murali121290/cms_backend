"""Export findings to Excel (multi-sheet workbook) and standalone HTML."""
from __future__ import annotations
import io
from collections import Counter
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
# Consistent font per xlsx skill guidance.
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid")  # emerald-800
BODY_FONT = Font(name="Arial", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
def _style_header(ws, row_idx: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
def _autofit(ws, max_widths: dict[int, int] | None = None) -> None:
    """Rough auto-width based on header + sampled cells. max_widths caps column widths."""
    max_widths = max_widths or {}
    for col_idx, col in enumerate(ws.columns, start=1):
        letter = get_column_letter(col_idx)
        length = 0
        for cell in col:
            if cell.value is None:
                continue
            val = str(cell.value)
            length = max(length, min(len(val), 80))
        width = min(max(length + 2, 10), max_widths.get(col_idx, 60))
        ws.column_dimensions[letter].width = width
def _fill_consistency_sheets(wb: Workbook, data: dict[str, Any], job_id: str) -> None:
    """Populate consistency sheets into an existing workbook."""
    # ---- Summary sheet ----
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Manuscript consistency report"
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    ws["A2"] = f"Job ID: {job_id}"
    ws["A2"].font = BODY_FONT
    row = 4
    ws.cell(row=row, column=1, value="Overview").font = Font(name="Arial", size=12, bold=True)
    row += 1
    for label, key in [
        ("Chapters", "chapter_count"),
        ("Total words (excl. excluded zones)", "total_words"),
        ("Total findings", "total_findings"),
        ("Inconsistencies", "total_inconsistencies"),
    ]:
        ws.cell(row=row, column=1, value=label).font = BODY_FONT
        ws.cell(row=row, column=2, value=data["meta"][key]).font = BODY_FONT
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="US/UK spelling profile").font = Font(name="Arial", size=12, bold=True)
    row += 1
    for label, val in [
        ("US-form occurrences", data["spelling_summary"]["us"]),
        ("UK-form occurrences", data["spelling_summary"]["uk"]),
        ("US share (%)", data["spelling_summary"]["us_percent"]),
        ("UK share (%)", data["spelling_summary"]["uk_percent"]),
    ]:
        ws.cell(row=row, column=1, value=label).font = BODY_FONT
        ws.cell(row=row, column=2, value=val).font = BODY_FONT
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="Findings by category").font = Font(name="Arial", size=12, bold=True)
    row += 1
    for cat, total in sorted(data["category_totals"].items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=1, value=cat).font = BODY_FONT
        ws.cell(row=row, column=2, value=total).font = BODY_FONT
        row += 1
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    # ---- Chapters sheet ----
    ws_c = wb.create_sheet("Chapters")
    headers = ["#", "Chapter", "Filename", "Words", "Segments", "Excluded paras", "Findings", "US", "UK"]
    ws_c.append(headers)
    _style_header(ws_c, 1, len(headers))
    for ch in data["chapters"]:
        sp = data["spelling_profile"].get(str(ch["index"]), {"US": 0, "UK": 0})
        ws_c.append([
            ch["index"],
            ch["name"],
            ch["filename"],
            ch["word_count"],
            ch["segment_count"],
            ch["excluded_paragraphs"],
            ch["finding_count"],
            sp.get("US", 0),
            sp.get("UK", 0),
        ])
    for row_cells in ws_c.iter_rows(min_row=2):
        for c in row_cells:
            c.font = BODY_FONT
    ws_c.freeze_panes = "A2"
    ws_c.auto_filter.ref = ws_c.dimensions
    _autofit(ws_c)
    # ---- Inconsistencies sheet ----
    ws_i = wb.create_sheet("Inconsistencies")
    headers = ["Category", "Rule label", "Canonical key", "Total count", "Chapters", "Variants (form Ã— count)"]
    ws_i.append(headers)
    _style_header(ws_i, 1, len(headers))
    for item in data["inconsistencies"]:
        variants = " | ".join(f"{v['form']} Ã— {v['count']}" for v in item["variants"])
        chapters = ", ".join(str(c) for c in item["chapters"])
        ws_i.append([
            item["category"],
            item["rule_label"],
            item["canonical"],
            item["total_count"],
            chapters,
            variants,
        ])
    for row_cells in ws_i.iter_rows(min_row=2):
        for c in row_cells:
            c.font = BODY_FONT
            c.alignment = WRAP
    ws_i.freeze_panes = "A2"
    ws_i.auto_filter.ref = ws_i.dimensions
    _autofit(ws_i, max_widths={6: 80, 2: 50})
    # ---- One sheet per category ----
    categories = ["te_point", "compound", "spelling", "bias", "article"]
    category_titles = {
        "te_point": "TE points",
        "compound": "Compounds",
        "spelling": "US-UK spelling",
        "bias": "Bias terms",
        "article": "Articles",
    }
    for cat in categories:
        rows = [f for f in data["findings"] if f["category"] == cat]
        if not rows:
            continue
        ws_cat = wb.create_sheet(category_titles[cat][:31])  # Excel 31-char limit
        headers = ["Chapter #", "Chapter", "Source", "Page", "Para", "Rule", "Surface", "Canonical", "Context"]
        ws_cat.append(headers)
        _style_header(ws_cat, 1, len(headers))
        # Sort by chapter then page then para for editor scanning order
        rows_sorted = sorted(
            rows,
            key=lambda r: (r["chapter_index"], r["page"], r["para_index"]),
        )
        for f in rows_sorted:
            ws_cat.append([
                f["chapter_index"],
                f["chapter_name"],
                f["source"],
                f["page"],
                f["para_index"],
                f["rule_label"],
                f["surface"],
                f["canonical"],
                # Clean out the ⟪⟫ markers for Excel — replace with [ ] for clarity
                f["context"].replace("⟪", "[").replace("⟫", "]"),
            ])
        for row_cells in ws_cat.iter_rows(min_row=2):
            for c in row_cells:
                c.font = BODY_FONT
        ws_cat.freeze_panes = "A2"
        ws_cat.auto_filter.ref = ws_cat.dimensions
        _autofit(ws_cat, max_widths={6: 40, 7: 25, 8: 25, 9: 80})
    # ---- All findings sheet (flat, for pivoting) ----
    ws_all = wb.create_sheet("All findings")
    headers = ["Chapter #", "Chapter", "Category", "Source", "Page", "Para", "Rule ID", "Rule label", "Surface", "Canonical", "Severity", "Context"]
    ws_all.append(headers)
    _style_header(ws_all, 1, len(headers))
    for f in sorted(data["findings"], key=lambda r: (r["chapter_index"], r["page"], r["para_index"])):
        ws_all.append([
            f["chapter_index"],
            f["chapter_name"],
            f["category"],
            f["source"],
            f["page"],
            f["para_index"],
            f["rule_id"],
            f["rule_label"],
            f["surface"],
            f["canonical"],
            f["severity"],
            f["context"].replace("⟪", "[").replace("⟫", "]"),
        ])
    for row_cells in ws_all.iter_rows(min_row=2):
        for c in row_cells:
            c.font = BODY_FONT
    ws_all.freeze_panes = "A2"
    ws_all.auto_filter.ref = ws_all.dimensions
    _autofit(ws_all, max_widths={7: 28, 8: 40, 9: 25, 10: 25, 12: 80})


def _fill_ia_sheet(wb: Workbook, data: dict[str, Any]) -> None:
    """Populate the IA Report sheet into an existing workbook."""
    ia = data.get("ia_report", {})
    rows = ia.get("rows", [])
    ch_indices = ia.get("chapter_indices", [])
    ch_names = ia.get("chapter_names", {})
    ws = wb.create_sheet("IA Report")
    # ---- Styles ----
    HDR_FONT  = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    HDR_FILL  = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid")
    GRP_FONT  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    GRP_FILL  = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
    _BODY_FONT = Font(name="Arial", size=10)
    HIT_FONT  = Font(name="Arial", size=10, bold=True, color="065F46")
    ZERO_FONT = Font(name="Arial", size=10, color="9CA3AF")
    ZERO_FILL = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    CTR = Alignment(horizontal="center", vertical="center")
    TOP = Alignment(vertical="top")
    total_cols = 3 + len(ch_indices) + 1  # A, B, C, Ch01…ChNN, Total
    # ---- Header row ----
    ch_labels = [ch_names.get(str(i), f"Ch{i:02d}") for i in ch_indices]
    headers = ["Elements", "Terms/Rules/Patterns", "Example"] + ch_labels + ["Total"]
    for col_idx, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CTR if col_idx > 3 else Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 18
    # ---- Data rows ----
    current_element: str | None = None
    element_start_row: int = 2
    data_row = 2
    for ia_row in rows:
        element  = ia_row["element"]
        pattern  = ia_row["pattern"]
        example  = ia_row["example"] or ""
        by_ch    = ia_row["by_chapter"]
        total    = ia_row["total"]
        has_data = total > 0
        if element != current_element:
            if current_element is not None and data_row - 1 >= element_start_row:
                if data_row - 1 > element_start_row:
                    ws.merge_cells(
                        start_row=element_start_row, start_column=1,
                        end_row=data_row - 1, end_column=1,
                    )
            current_element = element
            element_start_row = data_row
        if has_data:
            bfont = _BODY_FONT
            bfill = None
            nfont = HIT_FONT
        else:
            bfont = ZERO_FONT
            bfill = ZERO_FILL
            nfont = ZERO_FONT
        cell_a = ws.cell(row=data_row, column=1, value=element)
        cell_a.font = GRP_FONT
        cell_a.fill = GRP_FILL
        cell_a.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_b = ws.cell(row=data_row, column=2, value=pattern)
        cell_b.font = bfont
        cell_b.alignment = TOP
        if bfill:
            cell_b.fill = bfill
        cell_c = ws.cell(row=data_row, column=3, value=example)
        cell_c.font = bfont
        cell_c.alignment = TOP
        if bfill:
            cell_c.fill = bfill
        for ch_col, ch_idx in enumerate(ch_indices, start=4):
            count = by_ch.get(str(ch_idx), 0)
            cell = ws.cell(row=data_row, column=ch_col, value=count if has_data else None)
            cell.font = nfont if count else bfont
            cell.alignment = CTR
            if bfill and not count:
                cell.fill = bfill
        total_col = 4 + len(ch_indices)
        cell_t = ws.cell(row=data_row, column=total_col, value=total if total else None)
        cell_t.font = HIT_FONT if total else bfont
        cell_t.alignment = CTR
        if bfill and not total:
            cell_t.fill = bfill
        data_row += 1
    if current_element is not None and data_row - 1 >= element_start_row:
        if data_row - 1 > element_start_row:
            ws.merge_cells(
                start_row=element_start_row, start_column=1,
                end_row=data_row - 1, end_column=1,
            )
    # ---- Column widths ----
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 18
    for col_idx in range(4, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 8


def build_excel(data: dict[str, Any], job_id: str) -> bytes:
    """Return .xlsx bytes for the full analysis."""
    wb = Workbook()
    _fill_consistency_sheets(wb, data, job_id)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_combined_excel(data: dict[str, Any], job_id: str) -> bytes:
    """Return .xlsx bytes combining consistency report and IA report in one workbook."""
    wb = Workbook()
    _fill_consistency_sheets(wb, data, job_id)
    if data.get("ia_report"):
        _fill_ia_sheet(wb, data)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_ia_excel(data: dict[str, Any], job_id: str) -> bytes:
    """Return .xlsx bytes in the IA report format."""
    wb = Workbook()
    default_sheet = wb.active
    _fill_ia_sheet(wb, data)
    wb.remove(default_sheet)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_csv(data: dict[str, Any]) -> bytes:
    """Flat CSV of all findings. Matches the 'All findings' sheet."""
    import csv
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([
        "chapter_index", "chapter_name", "category", "source", "page", "para_index",
        "rule_id", "rule_label", "surface", "canonical", "severity", "context",
    ])
    for f in sorted(data["findings"], key=lambda r: (r["chapter_index"], r["page"], r["para_index"])):
        w.writerow([
            f["chapter_index"], f["chapter_name"], f["category"], f["source"],
            f["page"], f["para_index"], f["rule_id"], f["rule_label"],
            f["surface"], f["canonical"], f["severity"],
            f["context"].replace("⟪", "[").replace("⟫", "]"),
        ])
    return buf.getvalue().encode("utf-8-sig")  # BOM so Excel opens UTF-8 correctly
